import requests
from bs4 import BeautifulSoup
import re
import os # For os.path.exists and os.remove
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def scrape_phone_numbers(url):
    """
    Scrapes phone number links from a given URL.
    Returns a list of dictionaries, each with 'outer_html', 'text_content', 'href_value'.
    """
    try:
        response = requests.get(url, timeout=10)
        response.raise_for_status()
    except requests.exceptions.RequestException as e:
        print(f"Error fetching URL: {e}")
        return []

    if response.status_code == 200:
        soup = BeautifulSoup(response.content, 'html.parser')
        link_details = []
        for a_tag in soup.find_all('a', href=True):
            href_value = a_tag.get('href', '')
            if href_value.startswith('tel:'):
                link_details.append({
                    'outer_html': str(a_tag),
                    'text_content': a_tag.get_text(strip=True),
                    'href_value': href_value
                })
        return link_details
    else:
        print(f"Error: Received status code {response.status_code}")
        return []

def normalize_phone_number(phone_str):
    """
    Normalizes a phone number string.
    """
    if phone_str is None:
        return ""
    if phone_str.startswith('tel:'):
        phone_str = phone_str[4:]
    translation_table = str.maketrans(
        "０１２３４５６７８９－（）　",
        "0123456789-() "
    )
    phone_str = phone_str.translate(translation_table)
    normalized = re.sub(r'[-()\s]', '', phone_str)
    normalized = re.sub(r'[^\d]', '', normalized)
    return normalized

def export_to_excel(data_list, filename="phone_link_analysis.xlsx"):
    """
    Exports data to an Excel file with conditional formatting for mismatches.
    """
    output_dir = "output"
    # Ensure the output directory exists
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # The filename parameter is now treated as a base filename.
    # Prepend the output directory to it.
    full_filename = os.path.join(output_dir, filename)

    if not data_list:
        print(f"No data provided to export_to_excel function for {full_filename}.")
        return

    workbook = Workbook()
    sheet = workbook.active
    headers = ["outerHTML", "textContent", "href Value", "Normalized textContent", "Normalized href Value", "Status", "Matched Master Number"]
    sheet.append(headers)

    # Define fill styles
    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")     # Red for Critical Mistake
    yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid") # Yellow for Warning
    green_fill = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")  # Green for Pass
    # No specific fill for "N/A" status, will use default cell style

    for item in data_list:
        row_values = [
            item.get('outer_html', ''),
            item.get('text_content', ''),
            item.get('href_value', ''),
            item.get('normalized_text_content', ''),
            item.get('normalized_href_value', ''),
            item.get('status', ''), # Use the new 'status' field
            item.get('matched_master_number', None) # Use new 'matched_master_number'
        ]
        sheet.append(row_values)

        # Apply conditional formatting based on the 'status'
        current_row_status = item.get('status', '')
        fill_to_apply = None
        if current_row_status == "Critical Mistake":
            fill_to_apply = red_fill
        elif current_row_status == "Warning":
            fill_to_apply = yellow_fill
        elif current_row_status == "Pass":
            fill_to_apply = green_fill

        if fill_to_apply:
            for cell in sheet[sheet.max_row]:
                cell.fill = fill_to_apply
    try:
        workbook.save(full_filename)
        print(f"\nData successfully exported to {full_filename}")
    except Exception as e:
        print(f"Error saving Excel file '{full_filename}': {e}")

def test_scraper():
    """
    Tests core functionality: scraping, data structure, normalization, status assignment, and Excel export.
    """
    print("\n--- Running Automated Tests ---")
    test_url = "https://takasugi-test.blhomepage.info/"
    # Test Excel filename now includes the output directory
    test_excel_base_filename = "test_output_phone_analysis.xlsx"
    test_excel_filepath = os.path.join("output", test_excel_base_filename)

    # Define master numbers for this test run
    test_master_numbers = [
        normalize_phone_number("0123456789"), # For "Pass" and potential "Warning"
        normalize_phone_number("0989870038")  # For another "Pass"
    ]
    # Add a master number that is intentionally not found in hrefs to ensure "Critical Mistake" logic works
    # test_master_numbers.append(normalize_phone_number("9999999999")) # Not used if logic is href-first

    expected_total_links = 12 # Based on previous runs

    # Define expected outcomes for specific links from the test URL
    # Based on the `test_master_numbers` above.
    expected_results_for_test_cases = [
        {
            'href_value_to_find': "tel:0123456789", # Raw href to find the link
            'expected_text_content': "01-2345-6789", # Expected raw text
            'expected_norm_href': "0123456789",
            'expected_norm_text': "0123456789",
            'expected_status': "Pass",
            'expected_matched_master': "0123456789"
        },
        {
            'href_value_to_find': "tel:0989870038",
            'expected_text_content': "TEL:098‐987-0038",
            'expected_norm_href': "0989870038",
            'expected_norm_text': "0989870038",
            'expected_status': "Pass",
            'expected_matched_master': "0989870038"
        },
        {
            'href_value_to_find': "tel:090−9800−5776", # Not in master numbers
            'expected_text_content': "080-8516-3944",
            'expected_norm_href': "09098005776",
            'expected_norm_text': "08085163944",
            'expected_status': "Critical Mistake",
            'expected_matched_master': None
        },
        {
            'href_value_to_find': "tel:ここに番号", # Non-numeric, not in master
            'expected_text_content': "ここに番号",
            'expected_norm_href': "",
            'expected_norm_text': "",
            'expected_status': "Critical Mistake", # Or "N/A" if "" was a master, but it's not.
            'expected_matched_master': None
        },
        # Example for a potential "Warning" if data existed:
        # A link <a href="tel:0123456789">Call Us!</a> with "0123456789" as a master number.
        # norm_href="0123456789", norm_text="". Status: "Warning", matched_master="0123456789"
        # The current test site doesn't seem to have such a case easily.
        # For instance, the link with href="tel:01-2345-6789" (norm: 0123456789)
        # has text "お電話での問合せはこちらTEL：01-2345-6789" (norm: 0123456789), which is a "Pass".
        {
            'href_value_to_find': "tel:01-2345-6789", # This is another link on the page
            'expected_text_content': "お電話での問合せはこちらTEL：01-2345-6789",
            'expected_norm_href': "0123456789",
            'expected_norm_text': "0123456789",
            'expected_status': "Pass", # Because "0123456789" is a master number
            'expected_matched_master': "0123456789"
        }
    ]

    actual_raw_data = scrape_phone_numbers(test_url)
    test_processed_data = [] # This will store data with status, matching the main script's output structure

    if actual_raw_data:
        for scraped_link_info in actual_raw_data:
            data_item = {
                'outer_html': scraped_link_info.get('outer_html', ''),
                'text_content': scraped_link_info.get('text_content', ''),
                'href_value': scraped_link_info.get('href_value', ''),
                'normalized_text_content': normalize_phone_number(scraped_link_info.get('text_content')),
                'normalized_href_value': normalize_phone_number(scraped_link_info.get('href_value', '')),
                'status': "Critical Mistake", # Default
                'matched_master_number': None
            }
            norm_href = data_item['normalized_href_value']
            norm_text = data_item['normalized_text_content']

            if test_master_numbers: # Simulate using master numbers
                for master_num in test_master_numbers:
                    if norm_href == master_num:
                        data_item['matched_master_number'] = master_num
                        if norm_text == master_num:
                            data_item['status'] = "Pass"
                        else:
                            data_item['status'] = "Warning"
                        break
            else: # No master numbers provided for the test (should not happen with current definition)
                data_item['status'] = "N/A (No master numbers provided)"
            test_processed_data.append(data_item)

    failures = []
    if len(test_processed_data) != expected_total_links:
        failures.append(f"Expected total {expected_total_links} links, but processed {len(test_processed_data)} links.")

    for i, expected_case in enumerate(expected_results_for_test_cases):
        # Find the actual data item that corresponds to this test case
        # Matching based on the original href_value as it's the most stable identifier from the raw scrape
        actual_item_found = None
        for actual_item_candidate in test_processed_data:
            if actual_item_candidate['href_value'] == expected_case['href_value_to_find']:
                # Further check text_content if it's a distinguishing factor
                if 'expected_text_content' in expected_case and actual_item_candidate['text_content'] != expected_case['expected_text_content']:
                    continue # Not the specific item we're looking for if text differs
                actual_item_found = actual_item_candidate
                break

        if not actual_item_found:
            failures.append(f"Test Case {i+1} (href_to_find: {expected_case['href_value_to_find']}): Corresponding link not found in processed data.")
            continue

        # Assertions for the found item
        if actual_item_found['normalized_text_content'] != expected_case['expected_norm_text']:
            failures.append(f"Test Case {i+1}: Normalized text mismatch. Expected '{expected_case['expected_norm_text']}', Got '{actual_item_found['normalized_text_content']}'. For href '{expected_case['href_value_to_find']}'.")
        if actual_item_found['normalized_href_value'] != expected_case['expected_norm_href']:
            failures.append(f"Test Case {i+1}: Normalized href mismatch. Expected '{expected_case['expected_norm_href']}', Got '{actual_item_found['normalized_href_value']}'. For href '{expected_case['href_value_to_find']}'.")
        if actual_item_found['status'] != expected_case['expected_status']:
            failures.append(f"Test Case {i+1}: Status mismatch. Expected '{expected_case['expected_status']}', Got '{actual_item_found['status']}'. For href '{expected_case['href_value_to_find']}'.")
        if actual_item_found['matched_master_number'] != expected_case['expected_matched_master']:
            failures.append(f"Test Case {i+1}: Matched master number mismatch. Expected '{expected_case['expected_matched_master']}', Got '{actual_item_found['matched_master_number']}'. For href '{expected_case['href_value_to_find']}'.")

    # Test Excel file creation with the new data structure
    if not failures and test_processed_data:
        export_to_excel(test_processed_data, filename=test_excel_base_filename) # Pass base filename
        if not os.path.exists(test_excel_filepath):
            failures.append(f"Excel export failed: File '{test_excel_filepath}' not created.")
        else:
            try:
                os.remove(test_excel_filepath)
                print(f"Test Excel file '{test_excel_filepath}' created and removed successfully.")
            except OSError as e:
                failures.append(f"Error removing test Excel file '{test_excel_filepath}': {e}")
    elif not processed_actual_data and not failures:
        failures.append("No data was processed by test_scraper, so Excel export was not tested.")
    elif failures:
        print("Skipping Excel export in test_scraper due to earlier failures.")

    if not failures:
        print("All tests passed!")
    else:
        print("\n--- Test Failures ---")
        for failure in failures:
            print(f"- {failure}")
        print("-----------------------")

def get_master_phone_numbers():
    """
    Collects a list of 'master' phone numbers from the user.
    """
    master_numbers = []
    print("\n--- Master Phone Number Entry ---")
    print("Enter the correct 'master' phone numbers for this website/company, one per line.")
    print("Press Enter on an empty line to finish.")

    while True:
        try:
            raw_number = input("Enter master phone number (or press Enter to finish): ")
            if not raw_number.strip():
                break
            normalized = normalize_phone_number(raw_number)
            if normalized and normalized not in master_numbers:
                master_numbers.append(normalized)
                print(f"Added master number: {raw_number} (Normalized: {normalized})")
            elif not normalized: print(f"Skipped invalid input: {raw_number}")
            else: print(f"Skipped duplicate master number: {raw_number} (Normalized: {normalized})")
        except EOFError:
            print("EOF encountered while entering master numbers. Finishing entry.")
            break

    if master_numbers: print(f"Master phone numbers collected: {master_numbers}")
    else: print("No master phone numbers were entered.")
    print("---------------------------------")
    return master_numbers

if __name__ == "__main__":
    user_master_numbers = get_master_phone_numbers()

    if not user_master_numbers:
        print("No master phone numbers provided by user. Detailed comparison against master list will be skipped for the main scrape.")

    url = None
    try:
        url = input("Enter the URL to scrape: ")
    except EOFError:
        print("No input provided (EOFError). Proceeding without user URL for main scrape.")

    if url:
        link_data_list_main = scrape_phone_numbers(url)
        if link_data_list_main:
            print(f"\nFound {len(link_data_list_main)} phone link(s) from user URL. Processing for details and export...")
            processed_data_for_excel_main = []
            for scraped_link_info in link_data_list_main: # Renamed loop variable
                data_item = {
                    'outer_html': scraped_link_info.get('outer_html', ''),
                    'text_content': scraped_link_info.get('text_content', ''),
                    'href_value': scraped_link_info.get('href_value', ''),
                    'normalized_text_content': normalize_phone_number(scraped_link_info.get('text_content')),
                    'normalized_href_value': normalize_phone_number(scraped_link_info.get('href_value', '')),
                    'status': "Critical Mistake", # Default status
                    'matched_master_number': None
                }

                norm_href = data_item['normalized_href_value']
                norm_text = data_item['normalized_text_content']

                if user_master_numbers: # If master numbers were provided
                    for master_num in user_master_numbers:
                        if norm_href == master_num:
                            data_item['matched_master_number'] = master_num
                            if norm_text == master_num:
                                data_item['status'] = "Pass"
                            else:
                                data_item['status'] = "Warning"
                            break # Found a master match for href, no need to check other master numbers
                else: # No master numbers provided
                    data_item['status'] = "N/A (No master numbers provided)"
                    # Old is_mismatch logic for reference, but not directly used for status here
                    # data_item['is_mismatch_legacy'] = norm_text != norm_href

                processed_data_for_excel_main.append(data_item)

                print(f"  Outer HTML: {data_item['outer_html']}")
                print(f"  Text Content: {data_item['text_content']}")
                print(f"  HREF Value: {data_item['href_value']}")
                print(f"  Normalized Text: {data_item['normalized_text_content']}")
                print(f"  Normalized HREF: {data_item['normalized_href_value']}")
                print(f"  Status: {data_item['status']}")
                if data_item['matched_master_number']:
                    print(f"  Matched Master #: {data_item['matched_master_number']}")
                print("-" * 20)

            if processed_data_for_excel_main:
                export_to_excel(processed_data_for_excel_main) # Pass the main processed data directly
        else:
            print("No phone links found for the URL.")
            print("No data to export from user URL scrape.")
    else:
        print("No URL entered for main scrape. Exiting main scrape.")
        print("No data to export from user URL scrape.")

    test_scraper()
